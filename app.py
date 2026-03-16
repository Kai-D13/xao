import json
import math
import os
import time
import openpyxl
import requests
from flask import Flask, jsonify, render_template, Response

app = Flask(__name__)

SERVICES_KEY = "e3464a9335a846e985861bdf43fd8700201a93af28006a40"
TILEMAP_KEY  = "06fadcaa43886a1b8a3fd81709a1f9723bb3e25d1010554b"
EXCEL_PATH   = os.path.join(os.path.dirname(__file__), "check_failed.xlsx")
CACHE_PATH   = os.path.join(os.path.dirname(__file__), "geocache.json")

HUB_CARRIERS = {"NVCT Hub Di Linh - Lâm Đồng - Child"}

HUB_LOCATIONS = [
    {
        "name": "NVCT Hub Di Linh - Lâm Đồng - Child",
        "lat": 11.572849213854973,
        "lng": 108.04066512374126,
    },
]

HUB_DI_LINH_LAT = 11.572849213854973
HUB_DI_LINH_LNG = 108.04066512374126


# ── Distance ──────────────────────────────────────────────
def haversine_km(lat1, lng1, lat2, lng2):
    R = 6371
    dlat = math.radians(lat2 - lat1)
    dlng = math.radians(lng2 - lng1)
    a = (math.sin(dlat / 2) ** 2
         + math.cos(math.radians(lat1))
         * math.cos(math.radians(lat2))
         * math.sin(dlng / 2) ** 2)
    return round(2 * R * math.asin(math.sqrt(a)), 1)


# ── Cache ─────────────────────────────────────────────────
def load_cache():
    if os.path.exists(CACHE_PATH):
        with open(CACHE_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_cache(cache):
    try:
        with open(CACHE_PATH, "w", encoding="utf-8") as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
    except OSError:
        pass  # Read-only filesystem on Vercel — cache lives in memory only


# ── Geocode ───────────────────────────────────────────────
def geocode_address(address, cache):
    if address in cache:
        return cache[address]

    try:
        r1 = requests.get(
            "https://maps.vietmap.vn/api/search/v3",
            params={"apikey": SERVICES_KEY, "text": address},
            timeout=10,
        )
        r1.raise_for_status()
        results = r1.json()
        if not isinstance(results, list) or len(results) == 0:
            print(f"  [WARN] No search result: {address}")
            return None

        ref_id = results[0].get("ref_id")
        if not ref_id:
            return None

        r2 = requests.get(
            "https://maps.vietmap.vn/api/place/v3",
            params={"apikey": SERVICES_KEY, "refid": ref_id},
            timeout=10,
        )
        r2.raise_for_status()
        place = r2.json()
        lat = place.get("lat")
        lng = place.get("lng")

        if lat and lng:
            coord = {"lat": float(lat), "lng": float(lng)}
            cache[address] = coord
            save_cache(cache)
            return coord

    except Exception as e:
        print(f"  [ERROR] {address}: {e}")

    return None


# ── Load & deduplicate data ───────────────────────────────
def load_data():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    # Step 1: aggregate total_order_ward per (full_address, carrier_name)
    agg = {}
    for r in range(2, ws.max_row + 1):
        carrier      = ws.cell(r, 1).value
        province     = ws.cell(r, 3).value
        district     = ws.cell(r, 4).value
        ward         = ws.cell(r, 5).value
        total_order  = ws.cell(r, 6).value or 0
        full_address = ws.cell(r, 7).value

        if not full_address or not carrier:
            continue

        key = (full_address, carrier)
        if key not in agg:
            agg[key] = {
                "carrier_name": carrier,
                "full_address": full_address,
                "province_name": province,
                "district_name": district,
                "ward_name": ward,
                "total_order_sum": 0,
                "month_count": 0,
            }
        agg[key]["total_order_sum"] += float(total_order)
        agg[key]["month_count"] += 1

    # Step 2: for each full_address, keep the carrier with highest total_order_sum
    best = {}
    for (full_address, carrier), item in agg.items():
        if full_address not in best or item["total_order_sum"] > best[full_address]["total_order_sum"]:
            best[full_address] = item

    return list(best.values())


# ── Routes ────────────────────────────────────────────────
@app.route("/api/mapstyle")
def map_style():
    """Proxy VietMap style.json with patched glyphs URL."""
    r = requests.get(
        f"https://maps.vietmap.vn/api/maps/light/style.json?apikey={TILEMAP_KEY}",
        timeout=10,
    )
    style = r.json()
    style["glyphs"] = "https://fonts.openmaptiles.org/{fontstack}/{range}.pbf"
    return Response(
        json.dumps(style),
        content_type="application/json",
        headers={"Access-Control-Allow-Origin": "*"},
    )


@app.route("/")
def index():
    return render_template("index.html", tilemap_key=TILEMAP_KEY)


@app.route("/api/points")
def api_points():
    cache = load_cache()
    rows  = load_data()
    points = []

    for item in rows:
        address = item["full_address"]
        coord   = geocode_address(address, cache)
        if not coord:
            print(f"  [SKIP] {address}")
            continue

        avg_order = round(item["total_order_sum"] / item["month_count"], 1)
        is_hub    = item["carrier_name"] in HUB_CARRIERS
        distance  = haversine_km(
            HUB_DI_LINH_LAT, HUB_DI_LINH_LNG,
            coord["lat"], coord["lng"]
        )

        points.append({
            "lat":           coord["lat"],
            "lng":           coord["lng"],
            "carrier_name":  item["carrier_name"],
            "full_address":  address,
            "province_name": item["province_name"],
            "district_name": item["district_name"],
            "ward_name":     item["ward_name"],
            "avg_order":     avg_order,
            "total_order":   item["total_order_sum"],
            "is_hub_carrier": is_hub,
            "distance_km":   distance,
            "color":         "#2563EB" if is_hub else "#DC2626",
        })
        time.sleep(0.1)

    return jsonify({"points": points, "hubs": HUB_LOCATIONS})


@app.route("/api/geocode-status")
def geocode_status():
    cache = load_cache()
    rows  = load_data()
    total  = len(rows)
    cached = sum(1 for r in rows if r["full_address"] in cache)
    return jsonify({"total": total, "geocoded": cached, "pending": total - cached})


if __name__ == "__main__":
    print("Starting mapping server at http://localhost:5000")
    app.run(debug=True, port=5000)
